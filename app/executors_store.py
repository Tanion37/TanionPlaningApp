"""Список исполнителей (xlsx-вкладка executors)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .xlsx_store import default_xlsx_path

SHEET_NAME = "executors"
DEFAULT_EXECUTOR = "Юра"
DEFAULT_EXECUTORS: tuple[str, ...] = (DEFAULT_EXECUTOR, "Лёша", "Саша", "Сойер")

# Telegram @username для отправки списка в чат студии (без @).
# Личные ники – в config.json → executor_telegram, не в публичном коде.
EXECUTOR_TELEGRAM: dict[str, str] = {}


def normalize_executor_name(name: str | None) -> str:
    return (name or "").strip().casefold().replace("ё", "е")


def is_own_executor(name: str | None) -> bool:
    """Пустой исполнитель или Юра / Юрий — задача владельца, без отдельного раздела."""
    text = (name or "").strip()
    if not text:
        return True
    key = normalize_executor_name(text)
    return key in {"юра", "юрий"} or key.startswith("юрий ")


def telegram_username_for(name: str | None) -> str:
    import json

    key = normalize_executor_name(name)
    mapped = dict(EXECUTOR_TELEGRAM)
    try:
        from .paths import app_root

        cfg_path = app_root() / "config.json"
        if cfg_path.is_file():
            extra = json.loads(cfg_path.read_text(encoding="utf-8")).get("executor_telegram") or {}
            if isinstance(extra, dict):
                for raw_name, username in extra.items():
                    mapped[normalize_executor_name(str(raw_name))] = str(username).lstrip("@")
    except (OSError, json.JSONDecodeError, TypeError):
        pass
    return mapped.get(key, "")


class ExecutorsStore:
    def __init__(self, xlsx_path: Path | None = None) -> None:
        self.path = xlsx_path or default_xlsx_path()
        self.names: list[str] = []

    def load(self) -> list[str]:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            self.names = list(DEFAULT_EXECUTORS)
            self.save()
            return list(self.names)

        wb = load_workbook(self.path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self.names = list(DEFAULT_EXECUTORS)
            self.save()
            return list(self.names)

        ws = wb[SHEET_NAME]
        names: list[str] = []
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            val = row[0]
            if val is None or str(val).strip() == "":
                continue
            text = str(val).strip()
            if text.casefold() in {"исполнитель", "исполнители", "name", "имя"}:
                continue
            if text not in names:
                names.append(text)
        if not names:
            names = list(DEFAULT_EXECUTORS)
            self.names = names
            self.save()
            return list(self.names)
        if not any(n.casefold() == DEFAULT_EXECUTOR.casefold() for n in names):
            names = [DEFAULT_EXECUTOR, *names]
            self.names = names
            self.save()
            return list(self.names)
        self.names = names
        return list(self.names)

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            default = wb.active
            default.title = "tasks"
            from .xlsx_store import COLUMNS

            default.append(list(COLUMNS))

        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(["исполнитель"])
        for name in self.names:
            ws.append([name])
        wb.save(self.path)

    def add(self, name: str) -> tuple[bool, str]:
        text = (name or "").strip()
        if not text:
            return False, "Пустое имя."
        if any(n.casefold() == text.casefold() for n in self.names):
            return False, "already"
        self.names.append(text)
        self.save()
        return True, ""

    def remove(self, name: str) -> bool:
        key = (name or "").strip().casefold()
        before = len(self.names)
        self.names = [n for n in self.names if n.casefold() != key]
        if len(self.names) == before:
            return False
        self.save()
        return True
