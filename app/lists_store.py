"""Списки (xlsx-вкладка lists): колонки = списки, строка 1 = имя+алиасы."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

from .xlsx_store import default_xlsx_path

SHEET_NAME = "lists"
VIEW_SHEET = "lists_view"
DONE_PREFIX = "[x] "

# (каноническое имя, алиасы…)
DEFAULT_LISTS: tuple[tuple[str, ...], ...] = (
    ("Книги", "Почитать"),
    ("Фильмы", "Посмотреть", "Кино"),
    ("Игры", "Поиграть"),
    ("Покупки",),
    ("Идеи",),
)


@dataclass
class ListItem:
    text: str
    done: bool = False


@dataclass
class ListColumn:
    name: str
    aliases: list[str] = field(default_factory=list)
    items: list[ListItem] = field(default_factory=list)
    hide_done: bool = False

    def all_names(self) -> list[str]:
        return [self.name, *self.aliases]

    def header_cell(self) -> str:
        parts = [self.name, *self.aliases]
        return ",".join(parts)

    def visible_items(self) -> list[tuple[int, ListItem]]:
        out: list[tuple[int, ListItem]] = []
        for idx, item in enumerate(self.items):
            if self.hide_done and item.done:
                continue
            out.append((idx, item))
        return out


def parse_item_cell(raw: object) -> ListItem | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    done = False
    low = text.casefold()
    if low.startswith("[x]"):
        done = True
        text = text[3:].strip()
    elif text.startswith("☑"):
        done = True
        text = text[1:].strip()
    if not text:
        return None
    return ListItem(text=text, done=done)


def dump_item_cell(item: ListItem) -> str:
    if item.done:
        return DONE_PREFIX + item.text
    return item.text


class ListsStore:
    def __init__(self, xlsx_path: Path | None = None) -> None:
        self.path = xlsx_path or default_xlsx_path()
        self.columns: list[ListColumn] = []
        self.on_change: Callable[[], None] | None = None

    def load(self) -> list[ListColumn]:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            self.columns = self._defaults()
            self.save(notify=False)
            return list(self.columns)

        wb = load_workbook(self.path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self.columns = self._defaults()
            self.save(notify=False)
            return list(self.columns)

        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.columns = self._defaults()
            self.save(notify=False)
            return list(self.columns)

        hide_map = self._load_hide_map(wb)
        header = rows[0]
        cols: list[ListColumn] = []
        for col_idx, cell in enumerate(header):
            if cell is None or str(cell).strip() == "":
                continue
            parts = [p.strip() for p in str(cell).split(",") if p.strip()]
            if not parts:
                continue
            name, aliases = parts[0], parts[1:]
            items: list[ListItem] = []
            for row in rows[1:]:
                if col_idx >= len(row):
                    continue
                parsed = parse_item_cell(row[col_idx])
                if parsed is None:
                    continue
                items.append(parsed)
            cols.append(
                ListColumn(
                    name=name,
                    aliases=aliases,
                    items=items,
                    hide_done=bool(hide_map.get(name.casefold())),
                )
            )

        if not cols:
            cols = self._defaults()
            self.columns = cols
            self.save(notify=False)
            return list(self.columns)

        self.columns = cols
        return list(self.columns)

    def save(self, *, notify: bool = True) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            default = wb.active
            default.title = "tasks"
            from .xlsx_store import COLUMNS

            default.append(list(COLUMNS))

        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)

        if not self.columns:
            self.columns = self._defaults()

        max_items = max((len(c.items) for c in self.columns), default=0)
        headers = [c.header_cell() for c in self.columns]
        ws.append(headers)
        for i in range(max_items):
            row = []
            for c in self.columns:
                if i < len(c.items):
                    row.append(dump_item_cell(c.items[i]))
                else:
                    row.append(None)
            ws.append(row)

        if VIEW_SHEET in wb.sheetnames:
            del wb[VIEW_SHEET]
        view = wb.create_sheet(VIEW_SHEET)
        view.append(["name", "hide_done"])
        for col in self.columns:
            view.append([col.name, 1 if col.hide_done else 0])

        wb.save(self.path)
        if notify and self.on_change is not None:
            self.on_change()

    def _load_hide_map(self, wb) -> dict[str, bool]:
        if VIEW_SHEET not in wb.sheetnames:
            return {}
        out: dict[str, bool] = {}
        ws = wb[VIEW_SHEET]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or i == 0:
                continue
            name = str(row[0] or "").strip()
            if not name:
                continue
            flag = row[1] if len(row) > 1 else 0
            out[name.casefold()] = str(flag).strip() not in {"", "0", "false", "False", "нет"}
        return out

    def _defaults(self) -> list[ListColumn]:
        return [
            ListColumn(name=names[0], aliases=list(names[1:]), items=[])
            for names in DEFAULT_LISTS
        ]

    def resolve(self, needle: str) -> ListColumn | None:
        key = (needle or "").strip().casefold()
        if not key:
            return None
        for col in self.columns:
            for name in col.all_names():
                if name.casefold() == key:
                    return col
        hits = [
            col
            for col in self.columns
            if any(key in n.casefold() for n in col.all_names())
        ]
        if len(hits) == 1:
            return hits[0]
        return None

    def add_item(self, list_needle: str, item: str) -> tuple[ListColumn | None, str]:
        """Добавить пункт. Вернуть (колонка, сообщение об ошибке или '')."""
        text = (item or "").strip()
        if not text:
            return None, "Пустой пункт списка."
        col = self.resolve(list_needle)
        if col is None:
            return None, f"Список «{list_needle}» не найден."
        if any(existing.text == text for existing in col.items):
            return col, "already"
        col.items.append(ListItem(text=text, done=False))
        self.save()
        return col, ""

    def toggle_item(self, list_needle: str, index: int) -> ListColumn | None:
        col = self.resolve(list_needle)
        if col is None or index < 0 or index >= len(col.items):
            return None
        col.items[index].done = not col.items[index].done
        self.save()
        return col

    def set_hide_done(self, list_needle: str, hide: bool) -> ListColumn | None:
        col = self.resolve(list_needle)
        if col is None:
            return None
        col.hide_done = bool(hide)
        self.save()
        return col

    def snapshot(self) -> dict:
        return {
            "columns": [
                {
                    "name": col.name,
                    "aliases": list(col.aliases),
                    "hide_done": col.hide_done,
                    "items": [{"text": it.text, "done": it.done} for it in col.items],
                }
                for col in self.columns
            ]
        }

    def apply_snapshot(self, data: dict | None) -> None:
        raw_cols = (data or {}).get("columns") or []
        cols: list[ListColumn] = []
        for raw in raw_cols:
            name = str(raw.get("name") or "").strip()
            if not name:
                continue
            items: list[ListItem] = []
            for item in raw.get("items") or []:
                if isinstance(item, str):
                    parsed = parse_item_cell(item)
                    if parsed:
                        items.append(parsed)
                    continue
                text = str(item.get("text") or "").strip()
                if not text:
                    continue
                items.append(ListItem(text=text, done=bool(item.get("done"))))
            cols.append(
                ListColumn(
                    name=name,
                    aliases=[str(a).strip() for a in (raw.get("aliases") or []) if str(a).strip()],
                    items=items,
                    hide_done=bool(raw.get("hide_done")),
                )
            )
        if cols:
            self.columns = cols
            self.save(notify=False)
