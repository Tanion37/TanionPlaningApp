"""Чтение и запись задач в xlsx."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import Task, format_date, parse_date
from .projects import resolve_project_name, unify_project_casing
from .tags import (
    IMPORTANT_TAG,
    canonicalize_tag_key,
    collect_tag_usage,
    migrate_tag_list,
    normalize_tag_token,
    parse_tags_cell,
    tags_to_cell,
)


def _normalize_tag_list(tags: list[str], *, usage=None) -> list[str]:
    return migrate_tag_list(tags, usage=usage)


def _migrate_social_project(task: Task) -> bool:
    """Проект «соцсети» → тег соцсети, поле проекта очистить."""
    from .projects import project_key
    from .tags import SOCIAL_TAG

    if project_key(task.project) != "соцсети":
        return False
    task.project = ""
    if SOCIAL_TAG not in task.tags:
        task.tags.append(SOCIAL_TAG)
    return True

TZ = timezone(timedelta(hours=3))
TASKS_SHEET = "tasks"

COLUMNS: tuple[str, ...] = (
    "id",
    "название",
    "дата постановки",
    "дата фактического выполнения",
    "дата когда можно приступать",
    "дата когда надо закончить",
    "дата напоминания",
    "время напоминания",
    "периодичность напоминания",
    "кто поставил задачу",
    "исполнитель",
    "проект",
    "описание",
    "теги",
    "pos_x",
    "pos_y",
    "author_id",
    "chat_id",
    "source",
    "серия",
)

HEADER_ALIASES: dict[str, str] = {
    "id": "id",
    "название": "название",
    "title": "название",
    "дата постановки": "дата постановки",
    "created": "дата постановки",
    "дата фактического выполнения": "дата фактического выполнения",
    "completed": "дата фактического выполнения",
    "дата когда можно приступать": "дата когда можно приступать",
    "start": "дата когда можно приступать",
    "дата когда надо закончить": "дата когда надо закончить",
    "due": "дата когда надо закончить",
    "срок": "дата когда надо закончить",
    "дата напоминания": "дата напоминания",
    "remind": "дата напоминания",
    "время напоминания": "время напоминания",
    "remind_time": "время напоминания",
    "периодичность напоминания": "периодичность напоминания",
    "кто поставил задачу": "кто поставил задачу",
    "author": "кто поставил задачу",
    "исполнитель": "исполнитель",
    "executor": "исполнитель",
    "assignee": "исполнитель",
    "проект": "проект",
    "project": "проект",
    "описание": "описание",
    "description": "описание",
    "notes": "описание",
    "теги": "теги",
    "tags": "теги",
    "pos_x": "pos_x",
    "pos_y": "pos_y",
    "author_id": "author_id",
    "chat_id": "chat_id",
    "source": "source",
    "серия": "серия",
    "series": "серия",
    "series_id": "серия",
}


def default_xlsx_path(root: Path | None = None) -> Path:
    from .paths import app_root

    base = root or app_root()
    return base / "data" / "tasks.xlsx"


def _normalize_remind_time(value) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "hour") and hasattr(value, "minute"):
        # time / datetime from openpyxl
        try:
            return f"{int(value.hour):02d}:{int(value.minute):02d}"
        except (TypeError, ValueError):
            pass
    text = str(value).strip()
    if not text:
        return ""
    # "14:30" or "14:30:00" or ISO with T
    if "T" in text:
        text = text.split("T", 1)[1]
    text = text.replace(".", ":")
    parts = text.split(":")
    try:
        h, m = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
    except ValueError:
        return ""
    if h > 23 or m > 59:
        return ""
    return f"{h:02d}:{m:02d}"


def now_date() -> date:
    return datetime.now(TZ).date()


def _next_id(tasks: list[Task]) -> str:
    max_num = 0
    for task in tasks:
        match = re.fullmatch(r"(\d+)", str(task.id))
        if match:
            max_num = max(max_num, int(match.group(1)))
    return f"{max_num + 1:03d}"


def _header_map(row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        key = HEADER_ALIASES.get(str(cell).strip().lower())
        if key:
            mapping[key] = idx
    return mapping


def _cell(row: tuple, mapping: dict[str, int], name: str):
    idx = mapping.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


class TaskStore:
    def __init__(self, path: Path | None = None) -> None:
        self.path = path or default_xlsx_path()
        self.tasks: list[Task] = []

    def ensure_exists(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            return
        json_path = self.path.parent.parent / "tasks.json"
        if json_path.exists():
            self.tasks = self._from_json(json_path)
            self.save()
            return
        self.tasks = []
        self.save()

    def load(self) -> list[Task]:
        self.ensure_exists()
        # data_only=False: формулы не нужны; False надёжнее при свежем сохранении из Excel
        wb = load_workbook(self.path, data_only=False)
        # Важно: НЕ wb.active — после Excel активным может быть lists → «пустые» задачи
        if TASKS_SHEET in wb.sheetnames:
            ws = wb[TASKS_SHEET]
        else:
            ws = wb.active
            # если активный лист — не tasks (например lists), ищем по заголовку
            rows_probe = list(ws.iter_rows(values_only=True, max_row=1))
            header = rows_probe[0] if rows_probe else ()
            if "название" not in _header_map(tuple(header) if header else ()):
                for name in wb.sheetnames:
                    cand = wb[name]
                    probe = list(cand.iter_rows(values_only=True, max_row=1))
                    if probe and "название" in _header_map(tuple(probe[0])):
                        ws = cand
                        break
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.tasks = []
            return self.tasks

        mapping = _header_map(rows[0])
        if "название" not in mapping:
            # старый формат без заголовков – считаем первой колонкой название
            mapping = {"название": 0}

        tasks: list[Task] = []
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            title = _cell(row, mapping, "название")
            if title is None or str(title).strip() == "":
                continue
            tid = _cell(row, mapping, "id")
            tid_s = str(tid).strip() if tid not in (None, "") else ""
            author_id = _cell(row, mapping, "author_id")
            chat_id = _cell(row, mapping, "chat_id")
            pos_x = _cell(row, mapping, "pos_x")
            pos_y = _cell(row, mapping, "pos_y")
            remind_cell = _cell(row, mapping, "дата напоминания")
            remind_time_cell = _cell(row, mapping, "время напоминания")
            remind_time = _normalize_remind_time(remind_time_cell)
            if (
                not remind_time
                and hasattr(remind_cell, "hour")
                and (getattr(remind_cell, "hour", 0) or getattr(remind_cell, "minute", 0))
            ):
                remind_time = _normalize_remind_time(remind_cell)
            tasks.append(
                Task(
                    id=tid_s or _next_id(tasks),
                    title=str(title).strip(),
                    created_at=parse_date(_cell(row, mapping, "дата постановки")),
                    completed_at=parse_date(
                        _cell(row, mapping, "дата фактического выполнения")
                    ),
                    start_at=parse_date(
                        _cell(row, mapping, "дата когда можно приступать")
                    ),
                    due_at=parse_date(
                        _cell(row, mapping, "дата когда надо закончить")
                    ),
                    remind_at=parse_date(remind_cell),
                    remind_time=remind_time,
                    remind_period=str(
                        _cell(row, mapping, "периодичность напоминания") or ""
                    ).strip(),
                    author=str(
                        _cell(row, mapping, "кто поставил задачу") or ""
                    ).strip(),
                    executor=str(_cell(row, mapping, "исполнитель") or "").strip(),
                    project=str(_cell(row, mapping, "проект") or "").strip(),
                    description=str(_cell(row, mapping, "описание") or "").strip(),
                    tags=parse_tags_cell(str(_cell(row, mapping, "теги") or "")),
                    pos_x=float(pos_x) if pos_x not in (None, "") else None,
                    pos_y=float(pos_y) if pos_y not in (None, "") else None,
                    author_id=int(author_id) if author_id not in (None, "") else None,
                    chat_id=int(chat_id) if chat_id not in (None, "") else None,
                    source=str(_cell(row, mapping, "source") or "xlsx").strip(),
                    series_id=str(_cell(row, mapping, "серия") or "").strip(),
                )
            )
        usage = collect_tag_usage(t.tags for t in tasks)
        migrated = False
        for task in tasks:
            new_tags = _normalize_tag_list(task.tags, usage=usage)
            if new_tags != task.tags:
                task.tags = new_tags
                migrated = True
            if _migrate_social_project(task):
                migrated = True
        from .period_roll import ensure_series_ids

        if ensure_series_ids(tasks):
            migrated = True
        self.tasks = tasks
        if unify_project_casing(self.tasks) or migrated:
            self.save()
        else:
            # справочник тегов на отдельной вкладке
            self._ensure_tags_sheet()
        return self.tasks

    def _ensure_tags_sheet(self) -> None:
        from .tags_sheet import TAGS_SHEET, write_tags_sheet

        if not self.path.exists():
            return
        try:
            wb = load_workbook(self.path)
            if TAGS_SHEET not in wb.sheetnames:
                write_tags_sheet(wb)
                wb.save(self.path)
        except OSError:
            pass

    def save(self) -> None:
        """Перезаписать лист tasks, сохранив остальные вкладки (lists и т.д.)."""
        from .tags_sheet import write_tags_sheet

        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            wb = load_workbook(self.path)
            if TASKS_SHEET in wb.sheetnames:
                del wb[TASKS_SHEET]
            ws = wb.create_sheet(TASKS_SHEET, 0)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = TASKS_SHEET
        ws.append(list(COLUMNS))
        for task in self.tasks:
            ws.append(
                [
                    task.id,
                    task.title,
                    format_date(task.created_at),
                    format_date(task.completed_at),
                    format_date(task.start_at),
                    format_date(task.due_at),
                    format_date(task.remind_at),
                    getattr(task, "remind_time", "") or "",
                    task.remind_period,
                    task.author,
                    getattr(task, "executor", "") or "",
                    task.project,
                    task.description,
                    tags_to_cell(task.tags),
                    task.pos_x,
                    task.pos_y,
                    task.author_id,
                    task.chat_id,
                    task.source,
                    getattr(task, "series_id", "") or "",
                ]
            )
        write_tags_sheet(wb)
        wb.save(self.path)

    def add_task(self, title: str, **kwargs) -> Task:
        task = Task(
            id=_next_id(self.tasks),
            title=title.strip(),
            created_at=kwargs.get("created_at", now_date()),
            completed_at=kwargs.get("completed_at"),
            start_at=kwargs.get("start_at"),
            due_at=kwargs.get("due_at"),
            remind_at=kwargs.get("remind_at"),
            remind_time=str(kwargs.get("remind_time") or "").strip(),
            remind_period=kwargs.get("remind_period", ""),
            author=kwargs.get("author", ""),
            executor=str(kwargs.get("executor") or "").strip(),
            project=resolve_project_name(kwargs.get("project"), self.tasks),
            description=str(kwargs.get("description") or ""),
            tags=_normalize_tag_list(list(kwargs.get("tags") or [])),
            pos_x=kwargs.get("pos_x"),
            pos_y=kwargs.get("pos_y"),
            author_id=kwargs.get("author_id"),
            chat_id=kwargs.get("chat_id"),
            source=kwargs.get("source", "app"),
            series_id=str(kwargs.get("series_id") or "").strip(),
        )
        from .period_roll import ensure_task_series

        ensure_task_series(task)
        self.tasks.append(task)
        self.save()
        return task

    def get(self, task_id: str) -> Task | None:
        for task in self.tasks:
            if task.id == task_id:
                return task
        return None

    def remove_task(self, task_id: str) -> bool:
        before = len(self.tasks)
        self.tasks = [t for t in self.tasks if t.id != task_id]
        if len(self.tasks) == before:
            return False
        self.save()
        return True

    def insert_task(self, task: Task) -> Task:
        from .activity_log import apply_state_to_task, snapshot_dict

        existing = self.get(task.id)
        if existing is not None:
            apply_state_to_task(existing, snapshot_dict(task))
            self.save()
            return existing
        self.tasks.append(task)
        self.save()
        return task

    def _from_json(self, path: Path) -> list[Task]:
        data = json.loads(path.read_text(encoding="utf-8"))
        tasks: list[Task] = []
        for item in data.get("tasks", []):
            created = parse_date(item.get("created_at"))
            completed = parse_date(item.get("completed_at"))
            due = parse_date(item.get("due"))
            remind = parse_date(item.get("remind_at"))
            tags = list(item.get("tags") or [])
            # нормализуем текстовые теги из старого json
            from .tags import normalize_tag_token

            normalized: list[str] = []
            for tag in tags:
                key = normalize_tag_token(str(tag))
                if key and key not in normalized:
                    normalized.append(key)
            if item.get("status") == "done":
                from .tags import DONE_TAG

                if DONE_TAG not in normalized:
                    normalized.append(DONE_TAG)
            normalized = _normalize_tag_list(normalized)
            if item.get("priority") == "high" and IMPORTANT_TAG not in normalized:
                normalized.append(IMPORTANT_TAG)
            tasks.append(
                Task(
                    id=str(item.get("id", _next_id(tasks))),
                    title=str(item.get("title", "")).strip(),
                    created_at=created,
                    completed_at=completed,
                    start_at=None,
                    due_at=due,
                    remind_at=remind,
                    remind_time=_normalize_remind_time(
                        item.get("remind_time") or item.get("remind_at")
                    ),
                    remind_period="каждый день" if item.get("remind_daily") else "",
                    author=str(item.get("author") or ""),
                    executor=str(item.get("executor") or "").strip(),
                    project=str(item.get("project") or ""),
                    description=str(item.get("notes") or item.get("description") or ""),
                    tags=normalized,
                    author_id=item.get("author_id"),
                    chat_id=item.get("chat_id"),
                    source=str(item.get("source") or "json"),
                    series_id=str(item.get("series_id") or "").strip(),
                )
            )
        return tasks
